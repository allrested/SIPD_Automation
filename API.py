__author__ = 'tatsuya'
import configparser
import glob
import json
import os
import sys
import timeit
import xlsxwriter
from selenium import webdriver
from openpyxl import load_workbook
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from pandas import json_normalize

import LOGIN as login
import LOGOUT as logout

options = {
    'log-level':'error'
}
configur = configparser.ConfigParser()
#orig_stdout = sys.stdout
#f = open('info.API.txt', 'w+')
#sys.stdout = f
#Get semua file excel
try:
  folder = configur.get('api', 'folder')
except Exception:
  folder = "DATA_API"
file = glob.glob("{}/[!_][!~$]*.xlsx".format(folder))
#Fungsi Buka File Config
def write_file():
    configur.write(open('config.ini', 'w'))
#Fungsi Generate File Config
def set_file():
  configur['api'] = {'folder': folder,'output': 'OUTPUT_API', 'indexFileBegin': '0', 'limitFilePerFolder': len(file)}
  configur['api_files'] = {}
#Fungsi Generate Metadata dari File yang akan diproses
def set_file_index(index, mulai, batas, status):
  configur.set('api_files', "filename-{}".format(index), '{}'.format(file[index]))
  configur.set('api_files', "begin-{}".format(index), '{}'.format(mulai))
  configur.set('api_files', "start-{}".format(index), '1')
  configur.set('api_files', "limit-{}".format(index), '{}'.format(batas))
  configur.set('api_files', "complete-{}".format(index), status)
def set_session(status, session, currenturl):
  configur.set('data', "status", '{}'.format(status))
  configur.set('data', "session", '{}'.format(session))
  configur.set('data', "currenturl", '{}'.format(currenturl))

#Cek keberadaan file config
if not os.path.exists('config.ini'):
  set_file()
  write_file()
else:
  configur.read('config.ini')

#Definisi konfigurasi file yang akan diproses
try:
  begin = configur.getint('api', 'indexfilebegin')
  limit = configur.getint('api', 'limitfileperfolder')
  status = configur.getint('data', 'status')
  fout = configur.get('api', 'output')
except Exception:
  print("API Config Generated!")
  set_file()
  write_file()
  exit()

while status != 200:
  try:
    driver = webdriver.Chrome(ChromeDriverManager().install())
    baca = login.login(driver)
    print("Berhasil Login!\nKode: {}".format(baca))
    status = baca
    curl = driver.command_executor._url
    session_id = driver.session_id
    set_session(status, session_id, curl)
    write_file()
    exit()
  except Exception as err:
    #Jika terjadi kesalahan tampilkan diconsole dan lanjutkan
    driver.close()
    print("Gagal Login")
    status = 401
    set_session(status, "", "")
    write_file()
    continue

try:
  curl = configur.get('data', 'currenturl')
  session_id = configur.get('data', 'session')
  driver = webdriver.Remote(command_executor=curl,desired_capabilities={})
  driver.close()
  driver.session_id = session_id
  actionChains = ActionChains(driver)
except Exception as err:
  #Jika terjadi kesalahan tampilkan diconsole dan keluar
  print("Error")
  print(err)
  exit()

#Pembacaan file
for a in range(begin,limit):
  try:
    #Definisi nilai default dari metadata file yang akan diproses
    nama = "-"
    fbegin = 1
    flimit = 1
    process = "False"
    #Membuka file excel yang diproses
    workbook = load_workbook(file[a])
    worksheet = workbook['API']
    #Check metadata dari Config
    nama = configur.get('api_files', "filename-{}".format(a))
    fbegin = configur.getint('api_files', "begin-{}".format(a))
    flimit = configur.getint('api_files', "limit-{}".format(a))
    process = configur.get('api_files', "complete-{}".format(a))
    #print("posisi: {} namafile: {} process {}".format(a,nama,process))
    #Check Status pemrosesan file
    if(process.lower() == "false"):
      #Blok untuk memproses data
      #print("Informasi excel fbegin: {} flimit: {} process TRUE".format(fbegin,flimit))
      counter = 0
      #Membaca data dari file excel
      for b in range(fbegin,flimit):
        try:
          WebDriverWait(driver, 1000)
          url = ('https://bandung.sipd.kemendagri.go.id/daerah/main/plan/asmas/2022/tampil-verif-usulan/25/0?verif_skpd=0&valid_tapd=0')
          driver.execute_script('window.open("'+url+'")')
          driver.switch_to.window(driver.window_handles[-1])
          elem = driver.find_element(By.XPATH, "//*")
          json_content = driver.find_element(By.TAG_NAME, "pre").text
          json_data = json.loads(json_content)
          df = json_normalize(json_data, 'data', ['draw', 'recordsTotal', 'recordsFiltered'], 
                          record_prefix='data_')
          df.columns = df.columns.map(lambda x: x.split(".")[-1])
          df.to_excel("./{}/Usulan.xlsx".format(fout))
          configur.set('api_files', "begin-{}".format(a), '{}'.format(b))
          print("#{}. File Created: Usulan.xlsx".format(b))
          #print("Informasi kolom-{} col0: {} col1: {} col2: {}".format(b,col0,col1,col2))
          #Tracking progress dan disimpan ke file config
          counter = counter + 1
          fbegin = b
          set_file_index(a,fbegin,flimit,process)
          write_file()
        except Exception as err:
          #Jika terjadi kesalahan tampilkan diconsole dan lanjutkan
          print(err)
          pass
      #Cek status dari data yang diproses
      if(counter>0 or fbegin >= flimit):
        process = "True"
        set_file_index(a,fbegin,flimit,process)
        write_file()
        baca = logout.logout(driver)
        set_session(baca, "", "")
        print("Berhasil Logout!\nKode : {}".format(baca))
    elif(process.lower() == "true"):
      #Blok ketika data sudah diproses sebelumnya
      driver.get("https://bandung.sipd.kemendagri.go.id/daerah/main/plan/asmas/2022/25/0")
      driver.execute_script("rekomUsulan('11791','verif')")
      
      button = driver.find_element(By.ID, "s2id_autogen15")
      driver.implicitly_wait(10)
      ActionChains(driver).move_to_element(button).perform()
      driver.find_element(By.ID, "s2id_autogen15").click()
      driver.find_element(By.ID, "s2id_autogen16_search").send_keys("Belanja Transfer")
      driver.find_element(By.ID, "s2id_autogen16_search").send_keys(Keys.ENTER)
      
      driver.find_element(By.ID, "s2id_autogen17").click()
      driver.find_element(By.ID, "s2id_autogen18_search").send_keys("rekonstruksi jalan")
      driver.find_element(By.ID, "s2id_autogen18_search").send_keys(Keys.ENTER)

      driver.find_element(By.NAME, "rekom_teks").send_keys("Rekomendasi")
      driver.find_element(By.NAME, "rekom_volume").send_keys("10")
      driver.find_element(By.NAME, "rekom_satuan").send_keys("meter")
      driver.find_element(By.NAME, "rekom_anggaran").send_keys("10000")
      
      WebDriverWait(driver, 60).until(expected_conditions.presence_of_element_located((By.ID, "simpan_verval")))
      element = driver.find_element(By.ID, "simpan_verval")
      driver.execute_script("arguments[0].click();", element)
      counter = 0
      #print("Informasi excel fbegin: {} flimit: {} file sudah diproses (COMPLETE)".format(fbegin,flimit))
    else:
      #Blok ketika data tidak valid saat dijalankan sebelumnya
      counter = 0
      #print("Informasi excel fbegin: {} flimit: {} process ERROR {}".format(fbegin,flimit, process))
  except configparser.NoOptionError:
    #Blok ketika metadata belum tersedia di file config
    flimit = worksheet.max_row
    set_file_index(a,fbegin,flimit,process)
    print("Metadata Generated!")
    #print("Filename: {}\nBegin : {}\nLimit : {}\nProcess : {}".format(file[a],fbegin,flimit, process))
    continue
  except Exception as err:
    #Blok ketika terjadi kesalahan
    print(err)
    process = "Error"
    set_file_index(a,fbegin,flimit,process)
    break

write_file()
#sys.stdout = orig_stdout
#f.close()