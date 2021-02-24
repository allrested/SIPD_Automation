__author__ = 'tatsuya'
import configparser
import glob
import json
import os
import sys
import time
import xlrd
import xlsxwriter
import pandas as pd
import openpyxl
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.webdriver.support.ui import Select
from pandas import json_normalize

import LOGIN as login
import LOGOUT as logout

options = {
    'log-level':'error'
}
configur = configparser.ConfigParser()
#orig_stdout = sys.stdout
#f = open('info.API.txt', 'w+')
#sys.stdout = f
#Get semua file excel
try:
  folder = configur.get('api', 'folder')
except Exception:
  folder = "DATA_API"
file = glob.glob("{}/[!_][!~$]*.xlsx".format(folder))
#Fungsi Buka File Config
def write_file():
    configur.write(open('config.ini', 'w'))
#Fungsi Generate Metadata dari File yang akan diproses
def set_session(status, session, currenturl):
  configur.set('data', "status", '{}'.format(status))
  configur.set('data', "session", '{}'.format(session))
  configur.set('data', "currenturl", '{}'.format(currenturl))

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


#Cek keberadaan file config
if not os.path.exists('config.ini'):
  write_file()
else:
  configur.read('config.ini')

#Definisi konfigurasi file yang akan diproses
try:
  status = configur.getint('data', 'status')
except Exception:
  print("API Config Generated!")
  write_file()
  exit()

while status != 200:
  try:
    driver = webdriver.Chrome(ChromeDriverManager().install())
    baca = login.login(driver)
    print("Berhasil Login!\nKode: {}".format(baca))
    status = baca
    curl = driver.command_executor._url
    session_id = driver.session_id
    set_session(status, session_id, curl)
    write_file()
    exit()
  except Exception as err:
    #Jika terjadi kesalahan tampilkan diconsole dan lanjutkan
    driver.close()
    print("Gagal Login")
    status = 401
    set_session(status, "", "")
    write_file()
    continue

try:
  curl = configur.get('data', 'currenturl')
  session_id = configur.get('data', 'session')
  driver = webdriver.Remote(command_executor=curl,desired_capabilities={})
  driver.close()
  driver.session_id = session_id
  actionChains = ActionChains(driver)
except Exception as err:
  #Jika terjadi kesalahan tampilkan diconsole dan keluar
  print("Error")
  print(err)
  exit()

workbook = xlrd.open_workbook("./penyelia.xlsx")
worksheet = workbook.sheet_by_index(0)
fbegin = 1
'''
for a in range(fbegin,worksheet.nrows):
  try:
    WebDriverWait(driver, 1000)
    if worksheet.cell_type(a, 1) in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
      driver.get("https://bandung.sipd.kemendagri.go.id/daerah/main/budget/setup-user/2021/mitra/25/0")
      WebDriverWait(driver, 1000)
      nip = str(worksheet.cell(a, 2).value)
      nama = str(worksheet.cell(a, 3).value)
      username = str(worksheet.cell(a, 4).value)
      jabatan = str(worksheet.cell(a, 5).value)
      skpd1 = str(worksheet.cell(a, 6).value)
      skpd2 = str(worksheet.cell(a, 7).value)
      skpd3 = str(worksheet.cell(a, 8).value)
      skpd4 = str(worksheet.cell(a, 9).value)
      WebDriverWait(driver, 60).until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR, ".tambah-user")))
      element = driver.find_element(By.CSS_SELECTOR, ".tambah-user")
      driver.execute_script("arguments[0].click();", element)
      driver.find_element(By.NAME, "nip_user").send_keys(nip)
      driver.find_element(By.NAME, "nama_user").send_keys(nama)

      #element = driver.find_element(By.ID, "s2id_autogen1")
      #driver.execute_script("arguments[0].click();", element)
      #driver.find_element(By.ID, "s2id_autogen2_search").send_keys(jabatan)
      driver.find_element(By.NAME, "user_name").send_keys(username)
      print(nama)
      WebDriverWait(driver, 10000)
      print(nip)
      button = driver.find_element(By.CSS_SELECTOR, ".box-setup-skpd > input")
      driver.implicitly_wait(10)
      ActionChains(driver).move_to_element(button).perform()
      driver.execute_script("arguments[0].click();", button)
      button = driver.find_element(By.CSS_SELECTOR, ".box-spv-giat > input")
      driver.implicitly_wait(10)
      ActionChains(driver).move_to_element(button).perform()
      driver.execute_script("arguments[0].click();", button)
      WebDriverWait(driver, 1000)
      WebDriverWait(driver, 10).until(expected_conditions.presence_of_element_located((By.ID, "simpanuser")))
      element = driver.find_element(By.ID, "simpanuser")
      print("{} {}".format(a,nama))
      break
  except Exception as err:
    #Jika terjadi kesalahan tampilkan diconsole dan lanjutkan
    print(err)
    break
'''
try:
  WebDriverWait(driver, 1000)
  url = ("https://bandung.sipd.kemendagri.go.id/daerah/main/budget/setup-user/2021/mitra/tampil/25/0?start=0&length=-1")
  driver.execute_script('window.open("'+url+'")')
  driver.switch_to.window(driver.window_handles[-1])
  elem = driver.find_element(By.XPATH, "//*")
  json_content = driver.find_element(By.TAG_NAME, "pre").text
  json_data = json.loads(json_content)
  df = json_normalize(json_data, 'data', ['draw', 'recordsTotal', 'recordsFiltered'], 
                  record_prefix='data_')
  df.columns = df.columns.map(lambda x: x.split(".")[-1])
  df.to_excel("./PENYELIA/API_PENYELIA.xlsx")
  print("# File Created: API_PENYELIA.xlsx")
except Exception as err:
  #Jika terjadi kesalahan tampilkan diconsole dan lanjutkan
  print(err)
  pass

for a in range(fbegin,worksheet.nrows):
  try:
    WebDriverWait(driver, 1000)
    if worksheet.cell_type(a, 1) not in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
      driver.get("https://bandung.sipd.kemendagri.go.id/daerah/main/budget/setup-user/2021/mitra/25/0")
      WebDriverWait(driver, 1000)
      idx = int(worksheet.cell(a, 1).value)
      nip = str(worksheet.cell(a, 2).value)
      nama = str(worksheet.cell(a, 3).value)
      username = str(worksheet.cell(a, 4).value)
      jabatan = str(worksheet.cell(a, 5).value)
      skpd1 = str(worksheet.cell(a, 6).value)
      skpd2 = str(worksheet.cell(a, 7).value)
      skpd3 = str(worksheet.cell(a, 8).value)
      skpd4 = str(worksheet.cell(a, 9).value)
      driver.execute_script("setUserSKPD('{}');".format(idx))
      button = driver.find_element(By.ID, "tambah_skpd")
      driver.implicitly_wait(10)
      driver.execute_script("arguments[0].click();", button)
      if skpd1 not in [None,'']:
        driver.find_element(By.ID, "s2id_autogen3").send_keys(skpd1)
        driver.find_element(By.ID, "s2id_autogen3").send_keys(Keys.ENTER)
      if skpd2 not in [None,'']:
        driver.find_element(By.ID, "s2id_autogen3").send_keys(skpd2)
        driver.find_element(By.ID, "s2id_autogen3").send_keys(Keys.ENTER)
      if skpd3 not in [None,'']:
        driver.find_element(By.ID, "s2id_autogen3").send_keys(skpd3)
        driver.find_element(By.ID, "s2id_autogen3").send_keys(Keys.ENTER)
      if skpd4 not in [None,'']:
        driver.find_element(By.ID, "s2id_autogen3").send_keys(skpd4)
        driver.find_element(By.ID, "s2id_autogen3").send_keys(Keys.ENTER)
      WebDriverWait(driver, 1000)
      WebDriverWait(driver, 60).until(expected_conditions.presence_of_element_located((By.ID, "simpan_skpd")))
      button = driver.find_element(By.ID, "simpan_skpd")
      driver.execute_script("arguments[0].click();", button)
      WebDriverWait(driver, 1000)
      WebDriverWait(driver, 60).until(expected_conditions.presence_of_element_located((By.ID, "tambah_skpd")))
      button = driver.find_element(By.ID, "tambah_skpd")
      driver.execute_script("arguments[0].click();", button)
      WebDriverWait(driver, 1000)
      WebDriverWait(driver, 60).until(expected_conditions.presence_of_element_located((By.ID, "tutup_form_skpd")))
      button = driver.find_element(By.ID, "tutup_form_skpd")
      driver.execute_script("arguments[0].click();", button)
      print("{} {} SKPD Updated".format(a,nama))
  except Exception as err:
    #Jika terjadi kesalahan tampilkan diconsole dan lanjutkan
    print(err)
    pass

'''


workbook = xlrd.open_workbook("./PENYELIA/API_PENYELIA.xlsx")
worksheet = workbook.sheet_by_index(0)
fbegin = 1
for a in range(fbegin,worksheet.nrows):
  try:
    WebDriverWait(driver, 1000)
    idx = int(worksheet.cell(a, 1).value)
    url = ("https://bandung.sipd.kemendagri.go.id/daerah/main/budget/setup-user/2021/mitra/tampil-user-skpd/25/0?idxuser={}".format(idx))
    driver.execute_script('window.open("'+url+'")')
    driver.switch_to.window(driver.window_handles[-1])
    elem = driver.find_element(By.XPATH, "//*")
    json_content = driver.find_element(By.TAG_NAME, "pre").text
    json_data = json.loads(json_content)
    df = json_normalize(json_data, 'data', ['draw', 'recordsTotal', 'recordsFiltered'], 
                    record_prefix='data_')
    df.columns = df.columns.map(lambda x: x.split(".")[-1])
    append_df_to_excel("./PENYELIA/{}.xlsx".format(idx), df)
  except Exception as err:
    #Jika terjadi kesalahan tampilkan diconsole dan lanjutkan
    print(err)
    pass
'''
write_file()
#sys.stdout = orig_stdout
#f.close()