__author__ = 'tatsuya'
import configparser
import glob
import json
import os
import sys
import timeit
import xlsxwriter
import traceback
from selenium import webdriver
from openpyxl import load_workbook
from selenium.common.exceptions import TimeoutException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from pandas import json_normalize

import LOGIN as login
import LOGOUT as logout

idu = ''
usr = ''
options = {
    'log-level':'error'
}
configur = configparser.ConfigParser()
#orig_stdout = sys.stdout
#f = open('info.API.txt', 'w+')
#sys.stdout = f
#Get semua file excel
try:
  folder = configur.get('api', 'folder')
  username = configur.get('data', 'username')
  url = configur.get('data', 'url')
except Exception:
  folder = "DATA"
  username = "01.01"
  url = "bandung.sipd.kemendagri.go.id"
file = glob.glob("{}/[!_][!~$]*.xlsx".format(folder))
#Fungsi Buka File Config
def write_file():
    configur.write(open('config.ini', 'w'))
#Fungsi Generate File Config
def set_file():
  configur['api'] = {'folder': folder,'output': 'OUTPUT', 'indexFileBegin': '0', 'limitFilePerFolder': len(file)}
  configur['api_files'] = {}
#Fungsi Generate Metadata dari File yang akan diproses
def set_file_index(index, mulai, batas, status):
  configur.set('api_files', "filename-{}".format(index), '{}'.format(file[index]))
  configur.set('api_files', "begin-{}".format(index), '{}'.format(mulai))
  configur.set('api_files', "start-{}".format(index), '2')
  configur.set('api_files', "limit-{}".format(index), '{}'.format(batas))
  configur.set('api_files', "complete-{}".format(index), status)
def set_session(status, session, currenturl, idu, username):
  configur.set('data', "status", '{}'.format(status))
  configur.set('data', "session", '{}'.format(session))
  configur.set('data', "idu", '{}'.format(idu))
  configur.set('data', "username", '{}'.format(username))
  configur.set('data', "currenturl", '{}'.format(currenturl))

#Cek keberadaan file config
if not os.path.exists('config.ini'):
  set_file()
  write_file()
else:
  configur.read('config.ini')

#Definisi konfigurasi file yang akan diproses
try:
  begin = configur.getint('api', 'indexfilebegin')
  limit = configur.getint('api', 'limitfileperfolder')
  status = configur.getint('data', 'status')
  fout = configur.get('api', 'output')
except Exception:
  print("API Config Generated!")
  set_file()
  write_file()
  exit()

while status != 200:
  try:
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get("https://{}/daerah".format(url))
    print("Browser Terbuka!")
    curl = driver.command_executor._url
    session_id = driver.session_id
    username = configur.get('data', 'username')
    idu = configur.get('data', 'idu')
    set_session(200, session_id, curl, idu, username)
    write_file()
    exit()
  except Exception as err:
    #Jika terjadi kesalahan tampilkan diconsole dan lanjutkan
    driver.close()
    print("Gagal Masuk")
    status = 401
    set_session(status, "", "", idu, username)
    write_file()
    continue

try:
  curl = configur.get('data', 'currenturl')
  session_id = configur.get('data', 'session')
  driver = webdriver.Remote(command_executor=curl,desired_capabilities={})
  driver.close()
  driver.session_id = session_id
  actionChains = ActionChains(driver)
except Exception as err:
  #Jika terjadi kesalahan tampilkan diconsole dan keluar
  print("Browser tertutup!")
  #print(err)
  status = 401
  set_session(status, "", "", idu, username)
  write_file()
  exit()

#Pembacaan file
for a in range(begin,limit):
  try:
    #Definisi nilai default dari metadata file yang akan diproses
    nama = "-"
    fbegin = 2
    flimit = 1
    process = "False"
    #Membuka file excel yang diproses
    workbook = load_workbook(file[a])
    worksheet = workbook['Sheet1']
    #Check metadata dari Config
    nama = configur.get('api_files', "filename-{}".format(a))
    fbegin = configur.getint('api_files', "begin-{}".format(a))
    flimit = configur.getint('api_files', "limit-{}".format(a))
    process = configur.get('api_files', "complete-{}".format(a))
    #print("posisi: {} namafile: {} process {}".format(a,nama,process))
    #Check Status pemrosesan file
    if(process.lower() == "false"):
      #Blok untuk memproses data
      #print("Informasi excel fbegin: {} flimit: {} process TRUE".format(fbegin,flimit))
      counter = 0
      b=fbegin
      #Membaca data dari file excel
      while b < flimit:
        try:
          #driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
          usulanid = str(worksheet["B{}".format(b)].value)
          usulan = str(worksheet["G{}".format(b)].value)
          keterangan = str(worksheet['P{}'.format(b)].value)
          alamat = str(worksheet['R{}'.format(b)].value)
          user = str(worksheet['D{}'.format(b)].value)
          if not os.path.exists('config.ini'):
            set_file()
            write_file()
          else:
            configur.read('config.ini')  
          username = configur.get('data', 'username')
          password = configur.get('data', 'password')
          status = configur.get('data', 'status')

          if(user != username and status != 200):
            curl = driver.command_executor._url
            session_id = driver.session_id
            username = str(worksheet['D{}'.format(b)].value)
            password = configur.get('data', 'password')
            driver.get("https://bandung.sipd.kemendagri.go.id/daerah/logout?idu={}".format(idu))
            baca = login.login(driver, username, password)
            set_session(baca, session_id, curl, idu, username)
            write_file()
            print("Berhasil Login!\nKode: {}".format(baca))
            try:
              WebDriverWait(driver, 20).until(expected_conditions.presence_of_element_located((By.CSS_SELECTOR, ".dropdown-toggle > img")))
              driver.find_element(By.CSS_SELECTOR, ".dropdown-toggle > img").click()
              WebDriverWait(driver, 20).until(expected_conditions.presence_of_element_located((By.LINK_TEXT, "Logout")))
              keluar = driver.find_element(By.LINK_TEXT, "Logout")
              idu = keluar.get_attribute('href')
              idu = idu[idu.find('idu=')+4:]
              set_session(baca, session_id, curl, idu, username)
              write_file()
            except Exception:
              pass
            finally:
              print("Ganti User!\nuser : {}".format(username))
          else:
            b += 1
            url = ('https://bandung.sipd.kemendagri.go.id/daerah/main/plan/asmas/2022/25/0')
            driver.get(url)
            WebDriverWait(driver, 60).until(expected_conditions.presence_of_element_located((By.XPATH, "//button[@class=\'fcbtn btn btn-success btn-outline btn-1b windowtoggle tambah-usulan\']")))
            driver.find_element(By.XPATH, "//button[@class=\'fcbtn btn btn-success btn-outline btn-1b windowtoggle tambah-usulan\']").click()
            try:
                element_present = expected_conditions.presence_of_element_located((By.ID, 's2id_autogen3'))
                WebDriverWait(driver, 3).until(element_present)
                WebDriverWait(driver, 60).until(expected_conditions.element_to_be_clickable((By.ID, "s2id_autogen3")))
                option=driver.find_element(By.ID, "s2id_autogen3")
                option.click()
            except TimeoutException:
                print("Timed out waiting for page to load")
            finally:
                print("Page loaded") 
            #actionChains.move_to_element(option).perform()
            #action.move_to_element(option).click().perform()
            driver.find_element(By.ID, "s2id_autogen4_search").send_keys("Kabupaten/Kota")
            driver.find_element(By.ID, "s2id_autogen4_search").send_keys(Keys.ENTER)
            driver.find_element(By.NAME, "masalah_teks").send_keys(keterangan)
            option=driver.find_element(By.NAME, "alamat_teks")
            webdriver.ActionChains(driver).move_to_element(option).perform()
            WebDriverWait(driver, 60).until(expected_conditions.element_to_be_clickable((By.NAME, "alamat_teks")))
            driver.find_element(By.NAME, "alamat_teks").send_keys(alamat)
            option=driver.find_element(By.ID, "s2id_autogen5")
            webdriver.ActionChains(driver).move_to_element(option).perform()
            webdriver.ActionChains(driver).click_and_hold(option).perform()
            WebDriverWait(driver, 60).until(expected_conditions.element_to_be_clickable((By.ID, "s2id_autogen5")))
            driver.find_element(By.ID, "s2id_autogen6_search").send_keys(usulan)
            driver.find_element(By.ID, "s2id_autogen6_search").send_keys(Keys.ENTER)
            WebDriverWait(driver, 60).until(expected_conditions.presence_of_element_located((By.ID, "simpan_usulan")))
            element = driver.find_element(By.ID, "simpan_usulan")
            driver.execute_script("arguments[0].click();", element)
            print("#{}. Data executed: {}".format(b, usulanid))
            #Tracking progress dan disimpan ke file config
            counter = counter + 1
            fbegin = b
            set_file_index(a,fbegin,flimit,process)
            write_file()
        except Exception as err:
          #Jika terjadi kesalahan tampilkan diconsole dan lanjutkan
          print(err)
          print(traceback.format_exc())
          pass
      #Cek status dari data yang diproses
      if(counter>0 or fbegin >= flimit):
        process = "True"
        set_file_index(a,fbegin,flimit,process)
        write_file()
        #baca = logout.logout(driver)
        #set_session(baca, "", "")
        #print("Berhasil Logout!\nKode : {}".format(baca))
    elif(process.lower() == "true"):
      #Blok ketika data sudah diproses sebelumnya
      counter = 0
      #print("Informasi excel fbegin: {} flimit: {} file sudah diproses (COMPLETE)".format(fbegin,flimit))
    else:
      #Blok ketika data tidak valid saat dijalankan sebelumnya
      counter = 0
      #print("Informasi excel fbegin: {} flimit: {} process ERROR {}".format(fbegin,flimit, process))
  except configparser.NoOptionError:
    #Blok ketika metadata belum tersedia di file config
    flimit = worksheet.max_row+1
    set_file_index(a,fbegin,flimit,process)
    print("Metadata Generated!")
    #print("Filename: {}\nBegin : {}\nLimit : {}\nProcess : {}".format(file[a],fbegin,flimit, process))
    continue
  except Exception as err:
    #Blok ketika terjadi kesalahan
    print(err)
    process = "Error"
    set_file_index(a,fbegin,flimit,process)
    break

write_file()
#sys.stdout = orig_stdout
#f.close()